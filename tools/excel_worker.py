import openpyxl

def get_tikker(path):
    workbook = openpyxl.load_workbook(path)
    workbook._active_sheet_index = 1 # Выбрать номер листа с данными
    sheet_obj = workbook.active

    all_ticker = set()
    for index in range(2, sheet_obj.max_row + 1):
        ticker = sheet_obj.cell(index, 2).value
        all_ticker.add(ticker)

    return all_ticker

def price_writer(path, actual_price):
    workbook = openpyxl.load_workbook(path)
    workbook._active_sheet_index = 1 # Выбрать номер листа с данными
    sheet_obj = workbook.active

    for index in range(2, sheet_obj.max_row + 1):
        ticker = sheet_obj.cell(index, 2).value
        if actual_price[ticker] != None:
            sheet_obj.cell(index, 7).value = actual_price[ticker]
            sheet_obj.cell(index, 7).number_format = '#,##0.00 ₽'
            sheet_obj.cell(index, 7).fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7",fill_type = "solid")

    workbook.save(path)